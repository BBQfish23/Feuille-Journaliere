import sys
# Hack de compatibilité pour Python 3.13+ (Streamlit Cloud)
try:
    import imghdr
except ImportError:
    import types
    m = types.ModuleType("imghdr")
    m.what = lambda x, h=None: None
    sys.modules["imghdr"] = m

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import random
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURATION DE LA PAGE ---
st.set_page_config(page_title="Générateur d'Horaire Spa", layout="centered")

st.title("📅 Générateur d'Horaire Réception")
st.markdown("Générez la feuille journalière à partir d'un export **Deputy (CSV)** ou **Emprez (Excel hebdomadaire)**.")

# --- SÉLECTION DU THÈME ---
liste_themes = [
    "Standard", 
    "Printemps", "Été", "Automne", "Hiver",
    "Nouvel An", "Saint-Valentin", "Pâques", "Fête des Mères", "Fête des Pères", 
    "Saint-Jean-Baptiste", "Fête du Canada", "Action de grâce", "Halloween", "Temps des Fêtes",
    "Semaine de Relâche", "Jour de Pluie (Cocooning)", "Canicule", "Grosse Journée", "Événement VIP"
]

theme_choisi = st.selectbox("🎭 Choisissez un thème pour l'horaire :", liste_themes)

# Dictionnaire complet des thèmes
themes_config = {
    "Standard": {"emoji": "", "msg_fin": "BONNE JOURNÉE !", "font": "Calibri"},
    "Printemps": {"emoji": "🌱", "msg_fin": "BONNE JOURNÉE ! 🌷", "font": "Comic Sans MS"},
    "Été": {"emoji": "☀️", "msg_fin": "BONNE JOURNÉE SOUS LE SOLEIL ! 🕶️", "font": "Trebuchet MS"},
    "Automne": {"emoji": "🍂", "msg_fin": "BEL AUTOMNE ET BONNE JOURNÉE ! 🍁", "font": "Georgia"},
    "Hiver": {"emoji": "❄️", "msg_fin": "BONNE JOURNÉE ! ⛄", "font": "Century Gothic"},
    "Nouvel An": {"emoji": "🎉", "msg_fin": "BONNE ANNÉE ! 🥂", "font": "Trebuchet MS"},
    "Saint-Valentin": {"emoji": "🤍", "msg_fin": "JOYEUSE SAINT-VALENTIN ! 🕊️", "font": "Georgia"},
    "Pâques": {"emoji": "🐇", "msg_fin": "JOYEUSES PÂQUES ! 🥚", "font": "Comic Sans MS"},
    "Fête des Mères": {"emoji": "💐", "msg_fin": "BONNE FÊTE DES MÈRES ! 🌸", "font": "Georgia"},
    "Fête des Pères": {"emoji": "☕", "msg_fin": "BONNE FÊTE DES PÈRES ! 👔", "font": "Times New Roman"},
    "Saint-Jean-Baptiste": {"emoji": "⚜️", "msg_fin": "BONNE FÊTE NATIONALE ! ⚜️", "font": "Impact"},
    "Fête du Canada": {"emoji": "🍁", "msg_fin": "BONNE FÊTE DU CANADA ! 🎆", "font": "Impact"},
    "Action de grâce": {"emoji": "🦃", "msg_fin": "JOYEUSE ACTION DE GRÂCE ! 🍂", "font": "Georgia"},
    "Halloween": {"emoji": "👻", "msg_fin": "JOYEUSE HALLOWEEN ! 🎃", "font": "Impact"},
    "Temps des Fêtes": {"emoji": "🎄", "msg_fin": "JOYEUSES FÊTES ! ✨", "font": "Courier New"},
    "Semaine de Relâche": {"emoji": "⛷️", "msg_fin": "BONNE RELÂCHE ! ☕", "font": "Comic Sans MS"},
    "Jour de Pluie (Cocooning)": {"emoji": "🌧️", "msg_fin": "BONNE JOURNÉE COCOONING ! 🍵", "font": "Courier New"},
    "Canicule": {"emoji": "🌡️", "msg_fin": "RESTEZ AU FRAIS ! 🧊", "font": "Arial Black"},
    "Grosse Journée": {"emoji": "💪", "msg_fin": "EXCELLENTE JOURNÉE, ON LÂCHE PAS ! 🔥", "font": "Arial Black"},
    "Événement VIP": {"emoji": "⭐", "msg_fin": "EXCELLENTE JOURNÉE VIP ! 🥂", "font": "Times New Roman"}
}

theme_actuel = themes_config[theme_choisi]

# --- FONCTIONS UTILITAIRES ---
def extraire_prenom(nom_complet):
    if pd.isna(nom_complet) or nom_complet == '': return "N/A"
    return str(nom_complet).split(' ')[0]

def str_to_minutes(t_str):
    try:
        h, m = map(int, str(t_str).strip().split(':'))
        return h * 60 + m
    except:
        return 1440 

# --- CONSTANTES DATES ---
JOURS_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
MOIS_FR = ["janvier", "février", "mars", "avril", "mai", "juin", "juillet", "août",
           "septembre", "octobre", "novembre", "décembre"]

def formater_date_fr(date_obj):
    return f"{JOURS_FR[date_obj.weekday()]} {date_obj.day} {MOIS_FR[date_obj.month - 1]} {date_obj.year}"


# --- HELPERS POUR L'IMPORT EMPREZ (Excel hebdomadaire) ---

import unicodedata

def _normaliser(s):
    """minuscule + sans accents, pour comparer les prénoms de façon fiable."""
    s = unicodedata.normalize('NFD', str(s))
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower().strip()

# Liste de prénoms connus. Sert à détecter automatiquement les PRÉNOMS COMPOSÉS :
# dans Emprez le nom est "Nom Prénom", on prend le dernier mot comme prénom, MAIS
# si l'avant-dernier mot est lui aussi un prénom connu, le prénom est composé
# (ex. "Grigoratus Lorena Gabriela" -> "Lorena").
PRENOMS_CONNUS = {_normaliser(x) for x in [
    # Prénoms (et premiers éléments de prénoms composés) du personnel
    "Megan", "Mariela", "Leanne", "Lina", "Nicolas", "Louis", "Magalie", "Amery",
    "Yasmine", "Kymia", "Anais", "Anaïs", "Sarah", "Maude", "Adam", "Brithany",
    "Ariane", "Sabrina", "Lee", "Anne", "Marie", "Lyne", "Rosemarie", "Oceane",
    "Annabel", "Ana", "Gabrielle", "Gabriela", "Lorena", "Bethsaida", "Kianna",
    "Jasmine", "Alexandra", "Ophelie", "Alexane", "Michael", "Rosalie", "Maxime",
    "Cloe", "Joyce", "Christine", "Emily", "Naellie", "Shenna", "Laurie", "Elyse",
    "Alyssa", "Imen", "Elisabeth", "Chloe", "Brittany", "Rukhsar", "Janie", "Sofia",
    "Gary", "Christel", "Benjamin",
    # Prénoms courants additionnels
    "Marc", "Jean", "Pierre", "Paul", "Luc", "Eric", "David", "Julie", "Karine",
    "Stephanie", "Melanie", "Nathalie", "Isabelle", "Caroline", "Veronique",
    "Catherine", "Genevieve", "Audrey", "Camille", "Laurence", "Maude", "Noemie",
    "Laura", "Emma", "Olivia", "Charles", "William", "Thomas", "Samuel", "Antoine",
    "Felix", "Gabriel", "Raphael", "Mathis", "Nora", "Sara", "Lea", "Juliette",
    "Mia", "Alice", "Florence", "Maya", "Victoria", "Daphne", "Beatrice",
]}

# Filet de sécurité manuel : pour tout cas que la détection auto raterait, on force
# ici le prénom. Clé = nom de famille (1er mot dans Emprez, sans accents/minuscule),
# valeur = prénom à afficher. Exemple : "gonzales": "Ana"
PRENOMS_OVERRIDE = {
}

def extraire_prenom_emprez(nom_complet):
    """Détermine le prénom à afficher depuis un nom Emprez "Nom Prénom(s)"."""
    tokens = str(nom_complet).split()
    if not tokens:
        return "N/A"
    # 1) Override manuel prioritaire (par nom de famille = 1er mot)
    cle = _normaliser(tokens[0])
    if cle in PRENOMS_OVERRIDE:
        return PRENOMS_OVERRIDE[cle]
    # 2) Détection auto des prénoms composés : si l'avant-dernier mot est un
    #    prénom connu, c'est le début du prénom composé.
    if len(tokens) >= 3 and _normaliser(tokens[-2]) in PRENOMS_CONNUS:
        return tokens[-2]
    # 3) Par défaut : le dernier mot
    return tokens[-1]

def lire_excel(file):
    """Lit un fichier Excel Emprez peu importe le moteur (xlsx déguisé en .xls inclus)."""
    data = file.read()
    for eng in ('openpyxl', 'xlrd'):
        try:
            return pd.ExcelFile(io.BytesIO(data), engine=eng)
        except Exception:
            continue
    raise ValueError("Format Excel non reconnu.")

def parse_cellule_emprez(text):
    """Retourne (start, end, poste, note) ou None si la cellule ne contient pas de quart.

    Format d'une cellule Emprez :
        9h00 - 17h00
        Responsable Réception
        Principale
        F: ACCUEIL        <- commentaire (accueil / sur appel ...) sous la mention F:
    """
    t = str(text)
    if t.strip() == '' or t.strip().lower() == 'nan':
        return None
    m = re.search(r'(\d{1,2})h(\d{2})\s*-\s*(\d{1,2})h(\d{2})', t)
    if not m:
        return None
    start = f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
    end = f"{int(m.group(3)):02d}:{int(m.group(4)):02d}"
    # Le commentaire est désormais sous la mention "F:"
    fm = re.search(r'F\s*:\s*(.+)', t)
    note = fm.group(1).strip() if fm else ''
    # Le poste = tout le texte sauf la plage horaire et la ligne F:
    poste = re.sub(r'\d{1,2}h\d{2}\s*-\s*\d{1,2}h\d{2}', ' ', t)
    poste = re.sub(r'F\s*:.*', ' ', poste)
    poste = ' '.join(poste.split())
    return start, end, poste, note

def classifier_poste_emprez(poste):
    """Mappe un poste Emprez vers une 'Area' compatible avec la logique Deputy existante."""
    p = poste.lower()
    if 'maintenance' in p:
        return 'Maintenance- spa'
    if 'entretien' in p:
        return 'ENTRETIEN MÉNAGER Responsable'
    if 'soin' in p:
        return 'MASSO Responsable'
    if 'bistro' in p:
        return 'Bistro - Supervision'
    if 'opération' in p or 'operation' in p:
        return 'Site extérieur - Supervision'
    if 'superviseur' in p and ('réception' in p or 'reception' in p):
        return 'Réception - Supervision'
    if 'responsable réception' in p or 'responsable reception' in p:
        if 'lounge' in p:
            return 'Lounge'
        if 'qualité' in p or 'qualite' in p:
            return 'QUALITÉ ET BIEN ÊTRE'
        return 'RÉCEPTION- Responsable Principale'
    # Postes non gérés (réservations, ventes, stationnement...) : ne matchent aucun filtre
    return poste

def calc_total_emprez(start, end):
    """Durée du quart en heures (utilisé pour la logique des pauses)."""
    try:
        sh, sm = map(int, start.split(':'))
        eh, em = map(int, end.split(':'))
        return round(((eh * 60 + em) - (sh * 60 + sm)) / 60.0, 2)
    except Exception:
        return 8.0

def charger_emprez(df_raw, jour_col, date_str):
    """Construit un DataFrame au schéma Deputy à partir d'une colonne (journée) Emprez."""
    rows = []
    for idx in range(3, len(df_raw)):
        nom_complet = str(df_raw.iloc[idx, 0]).strip()
        if nom_complet == '' or nom_complet.lower() == 'nan':
            continue
        parsed = parse_cellule_emprez(df_raw.iloc[idx, jour_col])
        if parsed is None:
            continue
        start, end, poste, note = parsed
        area = classifier_poste_emprez(poste)
        # Dans Emprez le nom est "Nom Prénom" : on détecte le prénom (gère les
        # prénoms composés) puis on reconstruit "Prénom NomComplet" pour rester
        # compatible avec extraire_prenom() tout en gardant un identifiant unique.
        prenom = extraire_prenom_emprez(nom_complet)
        employee = f"{prenom} {nom_complet}".strip()
        rows.append({
            'Employee': employee,
            'Start Time': start,
            'End Time': end,
            'Area': area,
            'Note': note,
            'Total Time': calc_total_emprez(start, end),
            'Start Date': date_str,
        })
    return pd.DataFrame(
        rows,
        columns=['Employee', 'Start Time', 'End Time', 'Area', 'Note', 'Total Time', 'Start Date']
    )


# --- GÉNÉRATION DE LA FEUILLE JOURNALIÈRE (commun Deputy / Emprez) ---
def generer_horaire(df, date_obj, date_formatee):
    # --- CRÉATION DU EXCEL ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Horaire Spa"

    # Styles
    font_titre_secondaire = Font(color="000000", bold=True, size=16)
    font_superviseur_dept = Font(color="000000", bold=True, size=12)
    font_normal = Font(color="000000", size=11)
    font_normal_bold = Font(color="000000", bold=True, size=11)
    font_cercles = Font(size=24) 
    font_qbe = Font(color="FFFFFF", bold=True, size=11)

    dark_gray_fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
    light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    thin = Side(style='thin'); thick = Side(style='thick')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick_bottom = Border(left=thin, right=thin, top=thin, bottom=thick)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_no_wrap = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border_thick_all = Border(left=thick, right=thick, top=thick, bottom=thick)

    # 1. Titre Principal
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 45
    
    titre_texte = f"Département Réception | {date_formatee}"
    if theme_actuel["emoji"]:
        titre_texte = f"{theme_actuel['emoji']} {titre_texte} {theme_actuel['emoji']}"
        
    ws['A1'] = titre_texte
    ws['A1'].fill = dark_gray_fill
    police_titre_theme = Font(name=theme_actuel["font"], color="FFFFFF", bold=True, size=18)
    ws['A1'].font = police_titre_theme
    ws['A1'].alignment = center_align
    for c in range(1, 6): ws.cell(row=1, column=c).border = border_thick_all

    curr_row = 2

    # 2. Section SUPERVISEURS
    ws.row_dimensions[curr_row].height = 30
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=5)
    ws.cell(row=curr_row, column=1, value="SUPERVISEURS ET ADJOINTS").fill = light_gray_fill
    ws.cell(row=curr_row, column=1).font = font_titre_secondaire
    ws.cell(row=curr_row, column=1).alignment = center_align
    for c in range(1, 6): ws.cell(row=curr_row, column=c).border = border_thick_all
    curr_row += 1

    ordres_sup = [
        ("Réception", "Réception - Supervision"), ("Bistro", "Bistro - Supervision"), 
        ("Opérations", "Site extérieur - Supervision"), ("Entretien", "ENTRETIEN MÉNAGER"),
        ("Soins", "MASSO"), ("Maintenance", "Maintenance- spa")
    ]

    for label, search in ordres_sup:
        found = df[df['Area'].str.contains(search, case=False, na=False)]
        
        if label == "Maintenance":
            found = found[found['Note'].str.contains('Sur Appel', case=False, na=False)]
            if found.empty:
                found = pd.DataFrame([{'Employee': 'Adam', 'Start Time': 'Sur Appel', 'End Time': '', 'Total Time': 0}])
        else:
            found = found[(found['Area'].str.contains('Supervision|Responsable|Chef', case=False, na=False) | 
                           found['Note'].str.contains('Responsable|Supervision', case=False, na=False))]

        if not found.empty:
            start_m = curr_row
            items = found.sort_values('Start Time') if 'Start Time' in found.columns else found
            total_items = len(items) 
            
            for i, (_, r) in enumerate(items.iterrows()):
                is_first = (i == 0)
                is_last = (i == total_items - 1)
                
                cell_label = ws.cell(row=curr_row, column=1, value=label.upper())
                cell_label.font = font_superviseur_dept
                cell_label.border = Border(
                    left=thick, right=thick, 
                    top=thick if is_first else thin, bottom=thick if is_last else thin
                )
                
                ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=5)
                h_end = str(r['End Time'])
                h_info = f" ({r['Start Time']} - {h_end})" if h_end and h_end != '0' else f" ({r['Start Time']})"
                txt = f"{extraire_prenom(r['Employee'])}{h_info}"
                
                for c in range(2, 6):
                    cell = ws.cell(row=curr_row, column=c)
                    cell.font = font_normal
                    if c == 2: cell.value = txt
                    
                    cell.border = Border(
                        left=thick if c == 2 else thin, right=thick if c == 5 else thin, 
                        top=thick if is_first else thin, bottom=thick if is_last else thin
                    )
                curr_row += 1
                
            if total_items > 1:
                ws.merge_cells(start_row=start_m, start_column=1, end_row=curr_row-1, end_column=1)

    # --- PRÉPARATION GLOBALE DES CAISSES ET DES PAUSES ---
    rm = df[df['Area'].str.contains('RÉCEPTION- Responsable', case=False, na=False) & (df['Start Time'] < '12:00')].sort_values('Start Time')
    qbe = df[df['Area'].str.contains('QUALITÉ ET BIEN ÊTRE', case=False, na=False)].sort_values('Start Time')
    rs = df[df['Area'].str.contains('RÉCEPTION- Responsable', case=False, na=False) & (df['Start Time'] >= '12:00')].sort_values('Start Time')
    
    lg_data = df[df['Area'].str.contains('Lounge', case=False, na=False)].sort_values('Start Time')
    lg_list = list(lg_data.iterrows())
    total_lg = len(lg_list)

    # 1. Calcul des Caisses
    candidats_caisses = []
    for _, r in pd.concat([rm, rs]).iterrows():
        candidats_caisses.append((r['Employee'], str(r['Start Time']).strip()))
        
    for _, r in lg_list:
        note_str = str(r['Note']).lower()
        if 'accueil' not in note_str and 'acceuil' not in note_str:
            candidats_caisses.append((r['Employee'], str(r['Start Time']).strip()))

    candidats_caisses.sort(key=lambda x: str_to_minutes(x[1]))

    map_caisses = {}
    for idx, candidat in enumerate(candidats_caisses):
        map_caisses[candidat] = idx + 1

    # 2. Calcul des Pauses (Cerveau Global)
    candidats_pauses = []
    # Ajouter Réception et QBE
    for _, r in pd.concat([rm, rs, qbe]).iterrows():
        is_qbe = 'QUALITÉ ET BIEN ÊTRE' in str(r['Area']).upper()
        candidats_pauses.append((r['Employee'], str(r['Start Time']).strip(), r['Total Time'], is_qbe))
    # Ajouter Lounge
    for _, r in lg_list:
        candidats_pauses.append((r['Employee'], str(r['Start Time']).strip(), r['Total Time'], False))

    # Trier tout le monde strictement par heure d'arrivée
    candidats_pauses.sort(key=lambda x: str_to_minutes(x[1]))

    map_pauses = {}
    pauses_globales = set()
    interdits_globaux = ["18:00", "18:30"]
    interdits_qbe = ["10:00", "11:30", "13:00", "14:30", "16:00", "17:30", "19:00", "20:30"]

    # weekday(): lundi=0 ... dimanche=6  -> "la semaine" = lundi au jeudi (0 à 3)
    qbe_matin_1315 = date_obj.weekday() <= 3
    for emp, h_start, total_t, is_qbe in candidats_pauses:
        # QBE du matin (quart débutant avant midi) : pause de midi fixée à 13h15,
        # uniquement du lundi au jeudi. On ne réserve pas de créneau global
        # (13:15 ne tombe jamais sur un :00/:30).
        if qbe_matin_1315 and is_qbe and str_to_minutes(h_start) < 720:
            map_pauses[(emp, h_start)] = "13:15"
            continue
        try:
            val = float(total_t)
            if val < 5.5: 
                map_pauses[(emp, h_start)] = "15 min"
                continue
            
            if not h_start or ":" not in h_start:
                map_pauses[(emp, h_start)] = "-"
                continue
                
            fmt = "%H:%M"
            base = datetime.strptime(h_start, fmt) + timedelta(hours=3)
            earliest = datetime.strptime("11:30", fmt)
            current_pause = max(base, earliest)
            
            minute = current_pause.minute
            if 0 < minute <= 30:
                current_pause = current_pause.replace(minute=30)
            elif minute > 30:
                current_pause = current_pause + timedelta(hours=1)
                current_pause = current_pause.replace(minute=0)
                
            while True:
                str_pause = current_pause.strftime(fmt)
                if str_pause in interdits_globaux:
                    pass 
                elif is_qbe and str_pause in interdits_qbe:
                    pass 
                elif str_pause in pauses_globales:
                    pass 
                else:
                    pauses_globales.add(str_pause)
                    map_pauses[(emp, h_start)] = str_pause
                    break
                
                current_pause += timedelta(minutes=30)
                if current_pause.hour < 6: 
                    map_pauses[(emp, h_start)] = "-"
                    break
        except:
            map_pauses[(emp, h_start)] = "-"

    # 3. SECTIONS EMPLOYÉS
    sections_a_afficher = ["RÉCEPTION"]
    if not lg_data.empty: sections_a_afficher.append("LOUNGE")

    for section in sections_a_afficher:
        ws.row_dimensions[curr_row].height = 30
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=5)
        ws.cell(row=curr_row, column=1, value=section).fill = light_gray_fill
        ws.cell(row=curr_row, column=1).font = font_titre_secondaire
        for c in range(1, 6): ws.cell(row=curr_row, column=c).border = border_thick_all
        curr_row += 1
        
        headers = {1: "NOM", 2: "QUART / POSTE", 3: "Caisse", 4: "Lunch", 5: "TÂCHE / RESP."}
        for c in range(1, 6):
            cell = ws.cell(row=curr_row, column=c)
            cell.font = font_normal_bold
            cell.border = border_thin
            if c in headers: cell.value = headers[c]
        curr_row += 1

        if section == "RÉCEPTION":
            # --- LOGIQUE STRICTE DES TÂCHES SANS RÉPÉTITION ---
            taches_am = ["Nettoyage boutique", "Objet perdus", "Aide entre dept."]
            taches_pm = ["Objet perdus", "Aide entre dept."]
            random.shuffle(taches_am)
            random.shuffle(taches_pm)
            
            ordre_postes = ["Poste 4", "Poste 1", "Poste 3", "Poste 2", "Poste 5"]
            fin_postes = {p: "00:00" for p in ordre_postes} 

            for group, is_qbe in [(rm, False), (qbe, True), (rs, False)]:
                for i, (_, r) in enumerate(group.iterrows()):
                    ws.cell(row=curr_row, column=1, value=extraire_prenom(r['Employee']))
                    
                    h_end = str(r['End Time']).strip()
                    h_end = h_end if h_end != '0' and h_end != '' else ""
                    h_start = str(r['Start Time']).strip()
                    
                    if not is_qbe:
                        start_min = str_to_minutes(h_start)
                        postes_libres = [p for p in ordre_postes if str_to_minutes(fin_postes[p]) <= start_min]
                        
                        if postes_libres:
                            poste_attribue = postes_libres[0]
                        else:
                            poste_attribue = min(ordre_postes, key=lambda p: (str_to_minutes(fin_postes[p]), ordre_postes.index(p)))
                        
                        fin_postes[poste_attribue] = h_end if h_end else "23:59"
                        quart_txt = f"{h_start}-{h_end} / {poste_attribue}"
                        
                        # Assignation stricte de la tâche
                        is_am = start_min < 720 # 12:00
                        if poste_attribue == "Poste 1":
                            tache_finale = "Bye Bye Clés"
                        else:
                            if is_am:
                                tache_finale = taches_am.pop(0) if len(taches_am) > 0 else ""
                            else:
                                tache_finale = taches_pm.pop(0) if len(taches_pm) > 0 else ""
                    else:
                        quart_txt = f"{h_start}-{h_end}"
                        tache_finale = "QBE"
                        
                    ws.cell(row=curr_row, column=2, value=quart_txt).font = font_normal_bold
                    
                    # Récupération de la caisse
                    if is_qbe:
                        c_val = "VAM" if i == 0 else "VPM"
                    else:
                        c_val = map_caisses.get((r['Employee'], h_start), "")
                        
                    ws.cell(row=curr_row, column=3, value=c_val)
                    
                    # Récupération de la pause globale
                    pause_val = map_pauses.get((r['Employee'], h_start), "-")
                    ws.cell(row=curr_row, column=4, value=pause_val)
                    
                    ws.cell(row=curr_row, column=5, value=tache_finale)
                    
                    for c in range(1, 6):
                        cell = ws.cell(row=curr_row, column=c)
                        cell.border = border_thin
                        cell.font = font_normal
                        if is_qbe: 
                            cell.fill = dark_gray_fill
                            cell.font = font_qbe
                    curr_row += 1
            
            for c in range(1, 6): 
                ws.cell(row=curr_row-1, column=c).border = border_thick_bottom
                
        elif section == "LOUNGE":
            # 1. Trouver l'employé qui finit le plus tard pour assigner la Fermeture
            index_fermeture = -1
            max_end_minutes = -1
            
            for idx_temp, (_, r_temp) in enumerate(lg_list):
                end_mins = str_to_minutes(r_temp['End Time'])
                # L'utilisation de >= permet de donner la fermeture à celui qui a 
                # commencé le plus tard en cas d'égalité d'heure de fin.
                if end_mins >= max_end_minutes:
                    max_end_minutes = end_mins
                    index_fermeture = idx_temp

            # 2. Boucle d'affichage et d'assignation
            for i, (_, r) in enumerate(lg_list):
                ws.cell(row=curr_row, column=1, value=extraire_prenom(r['Employee']))
                
                h_end_lg = str(r['End Time']).strip()
                h_end_lg = h_end_lg if h_end_lg != '0' and h_end_lg != '' else ""
                h_start_lg = str(r['Start Time']).strip()
                
                ws.cell(row=curr_row, column=2, value=f"{h_start_lg}-{h_end_lg}").font = font_normal_bold
                
                # Assignation des tâches selon la nouvelle priorité
                note_str = str(r['Note']).lower()
                is_accueil = 'accueil' in note_str or 'acceuil' in note_str
                
                if is_accueil:
                    t = "Accueil"
                elif i == index_fermeture:
                    t = "Fermeture Lounge" # Donné à celui calculé par l'heure de fin
                elif i == 0:
                    t = "Ouverture Lounge"
                else:
                    t = ""
                    
                ws.cell(row=curr_row, column=5, value=t)
                
                cell_c = ws.cell(row=curr_row, column=3)
                if is_accueil: 
                    cell_c.fill = black_fill
                else: 
                    cell_c.value = map_caisses.get((r['Employee'], h_start_lg), "")
                    
                # Récupération de la pause globale
                pause_val_lg = map_pauses.get((r['Employee'], h_start_lg), "-")
                ws.cell(row=curr_row, column=4, value=pause_val_lg)
                
                for c in range(1, 6): 
                    ws.cell(row=curr_row, column=c).border = border_thin
                    ws.cell(row=curr_row, column=c).font = font_normal
                curr_row += 1
                
            for c in range(1, 6): 
                ws.cell(row=curr_row-1, column=c).border = border_thick_bottom

    # 4. SECTION FINALE
    cercles_18_colles = "○" * 18
    police_fin_theme = Font(name=theme_actuel["font"], color="000000", bold=True, size=16)
    
    sections_finales = [
        ("VENTES FIDÉLITÉS", cercles_18_colles, font_titre_secondaire), 
        ("SOINS", cercles_18_colles, font_titre_secondaire), 
        (theme_actuel["msg_fin"], None, police_fin_theme)
    ]
    
    for titre, contenu, police_a_utiliser in sections_finales:
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=5)
        cell_t = ws.cell(row=curr_row, column=1, value=titre)
        cell_t.fill = light_gray_fill
        cell_t.font = police_a_utiliser
        ws.row_dimensions[curr_row].height = 30
        
        for c in range(1, 6): ws.cell(row=curr_row, column=c).border = border_thick_all
        curr_row += 1
        
        if contenu:
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=5)
            cell_c = ws.cell(row=curr_row, column=1, value=contenu)
            cell_c.font = font_cercles
            cell_c.alignment = center_no_wrap
            ws.row_dimensions[curr_row].height = 45
            for c in range(1, 6): ws.cell(row=curr_row, column=c).border = border_thin
            curr_row += 1

   # --- AUTO-SIZE ET BORDURES FINALES ---
    max_row = curr_row - 1
    for r in range(1, max_row + 1):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            if not cell.alignment.horizontal: cell.alignment = center_align
            cb = cell.border
            cell.border = Border(left=(thick if c == 1 else cb.left), right=(thick if c == 5 else cb.right), 
                                top=(thick if r == 1 else cb.top), bottom=(thick if r == max_row else cb.bottom))

    # Ajustement des largeurs de colonnes
    for i in range(1, 6):
        col_letter = get_column_letter(i)
        
        if i == 1:
            ws.column_dimensions[col_letter].width = 25.5
        else:
            max_length = 0
            for row in ws.iter_rows(min_col=i, max_col=i):
                for cell in row:
                    if cell.coordinate not in ws.merged_cells or cell.column == 1:
                        try:
                            if cell.value:
                                length = len(str(cell.value))
                                if cell.font and cell.font.bold: length += 2
                                if length > max_length: max_length = length
                        except: pass
            ws.column_dimensions[col_letter].width = max_length + 3

    # --- RETOUR DES DONNÉES ---
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# --- INTERFACE PRINCIPALE : CHOIX DE LA SOURCE ---
source = st.radio(
    "📂 Source des données :",
    ["Deputy (CSV journalier)", "Emprez (Excel hebdomadaire)"]
)

df = None
date_obj = None
date_formatee = None

if source == "Deputy (CSV journalier)":
    uploaded_file = st.file_uploader("Choisir le fichier CSV", type="csv")
    if uploaded_file is not None:
        # 1. NETTOYAGE BLINDÉ DU FICHIER
        df = pd.read_csv(uploaded_file)
        df.columns = df.columns.str.strip()

        # FORCER les colonnes importantes à être du texte (string) pour éviter les crashs
        colonnes_texte = ['Start Time', 'End Time', 'Note', 'Area', 'Employee', 'Start Date']
        for col in colonnes_texte:
            if col in df.columns:
                df[col] = df[col].fillna('').astype(str).str.strip()
                df[col] = df[col].replace({'nan': '', '0.0': '', '0': ''})

        try:
            date_brute = df['Start Date'].iloc[0]
            date_obj = datetime.strptime(str(date_brute), '%Y-%m-%d')
            date_formatee = formater_date_fr(date_obj)
            st.success(f"Fichier détecté pour le : {date_formatee}")
        except Exception:
            st.error("Erreur avec la lecture de la date. Le fichier est-il valide ?")
            st.stop()

else:  # Emprez (Excel hebdomadaire)
    st.markdown(
        "Déposez l'export **Emprez** (Excel d'une semaine), puis choisissez la **journée** "
        "à générer (un seul jour est produit à la fois)."
    )
    uploaded_file = st.file_uploader("Choisir le fichier Excel", type=["xls", "xlsx"])
    if uploaded_file is not None:
        try:
            xls = lire_excel(uploaded_file)
            sheet_name = xls.sheet_names[0]
            df_raw = xls.parse(sheet_name, header=None)
        except Exception:
            st.error("Impossible de lire le fichier Excel Emprez.")
            st.stop()

        # Déterminer le lundi de la semaine (à partir du nom de l'onglet, ex. "... 2026-06-08 ...")
        m = re.search(r'(\d{4})-(\d{2})-(\d{2})', str(sheet_name))
        if m:
            lundi = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        else:
            today = datetime.today()
            lundi = today - timedelta(days=today.weekday())

        # Sélecteur de date : les 7 jours de la semaine (colonnes 1 à 7)
        jours_dispo = [lundi + timedelta(days=i) for i in range(7)]
        labels = [formater_date_fr(d) for d in jours_dispo]
        choix = st.selectbox("📅 Choisissez la journée à générer :", labels)
        idx_jour = labels.index(choix)
        date_obj = jours_dispo[idx_jour]
        date_formatee = labels[idx_jour]
        jour_col = idx_jour + 1  # colonne 0 = noms, colonne 1 = Lundi, etc.

        df = charger_emprez(df_raw, jour_col, date_obj.strftime('%Y-%m-%d'))
        if df.empty:
            st.warning("Aucun quart trouvé pour cette journée.")
            df = None
        else:
            st.success(f"{len(df)} employé(s) trouvé(s) pour le : {date_formatee}")

# --- TÉLÉCHARGEMENT ---
if df is not None and date_obj is not None:
    processed_data = generer_horaire(df, date_obj, date_formatee)
    st.download_button(
        label="📥 Télécharger l'horaire Excel",
        data=processed_data,
        file_name=f"Horaire_Spa_{date_obj.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
